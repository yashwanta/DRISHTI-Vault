"""Read the AMR Proxmox VM Tracker workbook and structure it for import.

Rules:
  * Site names, VM/application rows, network reference rows, and change log
    rows are parsed into plain (non-secret) metadata.
  * Any cell whose header is a "password"-like column and is non-empty is
    FLAGGED as a detected secret. It is NEVER auto-imported. The caller
    decides (with master-password confirmation) whether to import flagged
    secrets, one at a time.
  * No secret values are logged.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any

import openpyxl

PASSWORD_HEADERS = {"password", "passwd", "pwd", "secret", "token", "key", "pass"}


@dataclass
class ImportPreview:
    sites: list[dict] = field(default_factory=list)
    assets: list[dict] = field(default_factory=list)
    network: list[dict] = field(default_factory=list)
    changelog: list[dict] = field(default_factory=list)
    detected_secrets: list[dict] = field(default_factory=list)
    sheet_names: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "sites": self.sites,
            "assets": self.assets,
            "network": self.network,
            "changelog": self.changelog,
            "detected_secrets": self.detected_secrets,
            "detected_secrets_count": len(self.detected_secrets),
            "sheet_names": self.sheet_names,
        }


def _norm(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    # drop placeholder rows like "← Add new VM here"
    if s.startswith("←"):
        return ""
    return s


def _is_password_header(header: str) -> bool:
    h = (header or "").strip().lower()
    if not h:
        return False
    return any(p in h for p in PASSWORD_HEADERS)


def _detect_password_columns(header_row: list[str]) -> list[int]:
    return [i for i, h in enumerate(header_row) if _is_password_header(h)]


def parse_workbook(data: bytes) -> ImportPreview:
    """Parse an uploaded .xlsx file into a preview (no secrets imported)."""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    preview = ImportPreview(sheet_names=list(wb.sheetnames))

    seen_sites: set[str] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        name_lower = sheet_name.lower()

        if "sites index" in name_lower:
            _parse_sites_index(rows, preview, seen_sites)
        elif name_lower == "change log":
            _parse_change_log(rows, preview)
        elif name_lower == "network reference":
            _parse_network_reference(rows, preview)
        else:
            # Treat any other sheet as a per-site VM inventory (Springfield,
            # Hopkinsville, etc.). Derive the site name from the tab name.
            _parse_site_sheet(rows, preview, sheet_name, seen_sites)

    return preview


def _find_header_row(rows: list[tuple]) -> int:
    """Return the index of the row that most looks like a column header."""
    best, best_score = 0, 0
    for i, row in enumerate(rows[:6]):
        non_empty = sum(1 for c in row if _norm(c))
        if non_empty > best_score:
            best, best_score = i, non_empty
    return best


def _parse_sites_index(rows, preview: ImportPreview, seen_sites: set) -> None:
    header_idx = _find_header_row(rows)
    headers = [_norm(h).lower() for h in rows[header_idx]]
    for row in rows[header_idx + 1:]:
        cells = [_norm(c) for c in row]
        if not any(cells):
            continue
        name = _cell_by(headers, cells, ["site / plant", "site/plant", "site", "plant"])
        status = _cell_by(headers, cells, ["status"]) or "Active"
        notes = _cell_by(headers, cells, ["notes"])
        code = _cell_by(headers, cells, ["plant code", "code"])
        if name and name not in seen_sites:
            seen_sites.add(name)
            preview.sites.append({
                "name": name, "plant_code": code, "status": status,
                "notes": notes, "location": "",
            })


def _parse_change_log(rows, preview: ImportPreview) -> None:
    header_idx = _find_header_row(rows)
    headers = [_norm(h).lower() for h in rows[header_idx]]
    for row in rows[header_idx + 1:]:
        cells = [_norm(c) for c in row]
        if not any(cells):
            continue
        preview.changelog.append({
            "event_date": _cell_by(headers, cells, ["date"]),
            "asset_name": _cell_by(headers, cells, ["vm / application", "vm/application",
                                                    "asset", "application"]),
            "field_changed": _cell_by(headers, cells, ["field changed"]),
            "changed_by": _cell_by(headers, cells, ["changed by"]),
            "reason_ticket": _cell_by(headers, cells, ["reason / ticket #", "reason / ticket",
                                                       "reason", "ticket"]),
            "approved_by": _cell_by(headers, cells, ["approved by"]),
            "notes": "",
        })


def _parse_network_reference(rows, preview: ImportPreview) -> None:
    header_idx = _find_header_row(rows)
    headers = [_norm(h).lower() for h in rows[header_idx]]
    for row in rows[header_idx + 1:]:
        cells = [_norm(c) for c in row]
        if not any(cells):
            continue
        preview.network.append({
            "site_name": _cell_by(headers, cells, ["plant", "site"]),
            "vlan_id": _cell_by(headers, cells, ["vlan id", "vlan"]),
            "vlan_name": _cell_by(headers, cells, ["vlan name"]),
            "subnet": _cell_by(headers, cells, ["subnet"]),
            "gateway": _cell_by(headers, cells, ["gateway"]),
            "dhcp_scope": _cell_by(headers, cells, ["dhcp scope", "dhcp"]),
            "dns_servers": _cell_by(headers, cells, ["dns"]),
            "notes": _cell_by(headers, cells, ["notes"]),
        })


def _parse_site_sheet(rows, preview: ImportPreview, sheet_name: str,
                      seen_sites: set) -> None:
    header_idx = _find_header_row(rows)
    headers = [_norm(h).lower() for h in rows[header_idx]]
    pw_cols = _detect_password_columns(headers)
    site_name = sheet_name.replace("📋", "").strip()
    if site_name not in seen_sites:
        seen_sites.add(site_name)
        preview.sites.append({
            "name": site_name, "plant_code": "", "status": "Active",
            "notes": "Imported from workbook", "location": "",
        })

    for row in rows[header_idx + 1:]:
        cells = [_norm(c) for c in row]
        app = _cell_by(headers, cells, ["application / vm name", "app / vm name",
                                        "vm name", "application", "name"])
        if not app:
            continue
        asset = {
            "site_name": site_name,
            "app_vm_name": app,
            "asset_type": _guess_asset_type(app),
            "vm_id": _cell_by(headers, cells, ["vm id", "id"]),
            "hostname": "",
            "ip_address": _cell_by(headers, cells, ["ip address", "ip", "ipaddress"]),
            "environment": _cell_by(headers, cells, ["notes / environment", "environment"]),
            "web_url": _cell_by(headers, cells, ["web url", "url"]),
            "notes": "",
            "username": _cell_by(headers, cells, ["username", "user"]),
        }
        preview.assets.append(asset)

        # Detect (flag) any non-empty password cell — do NOT import value.
        for col in pw_cols:
            if col < len(cells):
                val = cells[col]
                if val:
                    preview.detected_secrets.append({
                        "site_name": site_name,
                        "app_vm_name": app,
                        "column_header": headers[col] if col < len(headers) else f"col{col}",
                        "field": "password",
                        "preview": "",  # never echo the value
                    })


def _guess_asset_type(name: str) -> str:
    n = (name or "").lower()
    if "proxmox host" in n or "pve" in n:
        return "Proxmox Host"
    if "aruba" in n or "ap-" in n:
        return "Aruba AP"
    if "switch" in n:
        return "Switch"
    if "shingocore" in n:
        return "ShingoCore"
    if "fleet" in n:
        return "FleetManager"
    if "rds" in n:
        return "RDS Core"
    if "postgres" in n or "database" in n or "db" in n:
        return "Database"
    if "core" in n:
        return "Ubuntu Server"
    return "Ubuntu Server"


def _cell_by(headers: list[str], cells: list[str], candidates: list[str]) -> str:
    for cand in candidates:
        for i, h in enumerate(headers):
            if h == cand or (cand and cand in h):
                if i < len(cells):
                    return cells[i]
    return ""
